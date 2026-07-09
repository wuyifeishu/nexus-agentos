"""
xlsx — Excel (.xlsx) operations using openpyxl.

Actions: read, headers, sheets, to_csv, stats, search
"""

from typing import Any


def run(action: str = "read", file_path: str = "", sheet: str = "", **kwargs: Any) -> str:
    try:
        import openpyxl
    except ImportError:
        return "[xlsx] openpyxl not installed. Run: pip install openpyxl"

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        return f"[xlsx] File not found: {file_path}"
    except Exception as e:
        return f"[xlsx] Error: {e}"

    ws = wb[sheet] if sheet else wb.active

    if action == "sheets":
        return "Sheets: " + ", ".join(wb.sheetnames)

    if action == "headers":
        headers = [cell.value for cell in ws[1]]
        return f"Headers ({len(headers)}): " + ", ".join(str(h) for h in headers if h)

    if action == "stats":
        rows = ws.max_row - 1
        cols = ws.max_column
        return f"Sheet: {ws.title}, Rows: {rows}, Columns: {cols}"

    if action == "read":
        lines = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True):
            lines.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(lines[:50])

    if action == "to_csv":
        rows_list = []
        for row in ws.iter_rows(values_only=True):
            rows_list.append(",".join(f'"{c}"' if c is not None else "" for c in row))
        return "\n".join(rows_list[:500])

    return f"[xlsx] Unknown action: {action}"


__all__ = ["run"]
